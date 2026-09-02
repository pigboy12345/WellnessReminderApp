"""Loads and validates today's reminders from WellnessReminders.xlsx."""
import datetime as dt
import logging
from typing import List, Tuple

import openpyxl

from src.models import Reminder

REQUIRED_FIELDS = ["id", "title", "message", "schedulingTime", "enabled"]

logger = logging.getLogger("WellnessReminder")


def _parse_time(value) -> dt.time:
    """Accepts a datetime.time, datetime.datetime, or 'HH:MM[:SS]' string."""
    if isinstance(value, dt.datetime):
        return value.time()
    if isinstance(value, dt.time):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return dt.datetime.strptime(value, fmt).time()
            except ValueError:
                continue
    raise ValueError(f"Unrecognized time format: {value!r}")


def _parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in ("true", "1", "yes", "y")
    return False


def load_todays_reminders(excel_path: str, today: dt.date = None) -> Tuple[List[Reminder], List[dict]]: # type: ignore
    today = today or dt.date.today()
    sheet_name = today.strftime("%d-%m-%Y")

    valid: List[Reminder] = []
    invalid: List[dict] = []

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except FileNotFoundError:
        logger.error("Excel file not found: %s", excel_path)
        return valid, invalid
    except Exception as exc:
        logger.error("Failed to open workbook '%s': %s", excel_path, exc)
        return valid, invalid

    if sheet_name not in wb.sheetnames:
        logger.warning("No sheet found for today's date '%s' in %s", sheet_name, excel_path)
        # Try fallback to a "Default" sheet
        default_sheet = "Default"
        if default_sheet in wb.sheetnames:
            sheet_name = default_sheet
            logger.info("Falling back to default sheet '%s'", default_sheet)
        else:
            logger.warning("No default sheet '%s' found either. Returning empty reminders.", default_sheet)
            wb.close()
            return valid, invalid

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        logger.warning("Sheet '%s' is empty.", sheet_name)
        wb.close()
        return valid, invalid

    header_cells = rows[0]
    headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
    header_index = {h: i for i, h in enumerate(headers)}

    missing_headers = [h for h in REQUIRED_FIELDS if h not in header_index]
    if missing_headers:
        logger.error("Sheet '%s' is missing required headers: %s", sheet_name, missing_headers)
        wb.close()
        return valid, invalid

    def cell(row, name):
        idx = header_index.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx].value

    for row_num, row in enumerate(rows[1:], start=2):
        if all(c.value is None for c in row):
            continue  # blank row, skip silently

        raw = {h: cell(row, h) for h in headers if h}

        missing = [f for f in REQUIRED_FIELDS if raw.get(f) in (None, "")]
        if missing:
            reason = f"Missing required field(s): {', '.join(missing)}"
            logger.warning("Row %d skipped (sheet '%s'): %s | data=%s", row_num, sheet_name, reason, raw)
            invalid.append({"row": row_num, "reason": reason, "data": raw})
            continue

        try:
            scheduling_time = _parse_time(raw.get("schedulingTime"))
        except ValueError as exc:
            reason = f"Invalid schedulingTime: {exc}"
            logger.warning("Row %d skipped (sheet '%s'): %s", row_num, sheet_name, reason)
            invalid.append({"row": row_num, "reason": reason, "data": raw})
            continue

        enabled = _parse_bool(raw.get("enabled"))
        if not enabled:
            continue  # not an error - simply excluded per spec

        icon = raw.get("icon/image")
        category = raw.get("category") or "General"

        try:
            reminder = Reminder(
                id=str(raw["id"]).strip(),
                title=str(raw["title"]).strip(),
                message=str(raw["message"]).strip(),
                category=str(category).strip(),
                scheduling_time=scheduling_time,
                enabled=True,
                icon=str(icon).strip() if icon else None,
            )
        except Exception as exc:
            reason = f"Unexpected error building reminder: {exc}"
            logger.warning("Row %d skipped (sheet '%s'): %s", row_num, sheet_name, reason)
            invalid.append({"row": row_num, "reason": reason, "data": raw})
            continue

        valid.append(reminder)

    wb.close()
    valid.sort(key=lambda r: r.scheduling_time)
    logger.info(
        "Loaded %d valid reminder(s) and %d invalid row(s) from sheet '%s'.",
        len(valid), len(invalid), sheet_name,
    )
    return valid, invalid
